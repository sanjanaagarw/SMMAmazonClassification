#!/usr/bin/env python
# -*- coding: utf-8 -*-
#######################

####################################
import argparse
import codecs
import time
import sys
import os, re, glob
import nltk
from collections import defaultdict
from random import shuffle, randint
import numpy as np
from numpy import array, arange, zeros, hstack, argsort
import unicodedata
from scipy.sparse import csr_matrix
# sklearn imports
import string
from random import shuffle
from random import sample
from sklearn.svm import SVC
from sklearn import preprocessing
from sklearn.cross_validation import StratifiedKFold
from sklearn.grid_search import GridSearchCV
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from sklearn import metrics
from sklearn.cross_validation import train_test_split
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_selection import SelectKBest, f_classif, chi2
from sklearn.multiclass import OneVsOneClassifier, OneVsRestClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn import cross_validation
from sklearn.preprocessing import LabelBinarizer
from openpyxl import load_workbook
n_jobs = 25


####################################
#fetch like getSAIMAThreeColumnFormat()---->getmyxls

def getmyxls():
    #"Girl on the train(done).xlsx", "Lights we cannot see (done).xlsx","NewHH.xlsx","Rogue Lawyer(done).xlsx","Secret Garden (done).xlsx","Set a watchman (done).xlsx"
    #listofxlsx=["Girl on the train(done).xlsx", "Lights we cannot see (done).xlsx","NewHH.xlsx","Rogue Lawyer(done).xlsx","Secret Garden (done).xlsx","Set a watchman (done).xlsx"]
    #listofxlsx=["Girl on the train(done).xlsx","Rogue Lawyer(done).xlsx"]
    #listofxlsxphones=["BLUStudio.xlsx","FireMobilePhone.xlsx","HTCDesire.xlsx","iPhone6.xlsx","MotorolaNexus6.xlsx","NokiaLumia.xlsx"]
    listofxlsxlaptops=["AcerChromebook.xlsx","AsusX551.xlsx","HPStream.xlsx"]
    listwetrulywant=[]
    for xl in listofxlsxlaptops:
        print "xlsx=", xl
        wb = load_workbook(filename = xl, use_iterators=True,data_only=True) #Name of the workbook. read_only to be used,
        sheetName = wb['Sheet2']
        fulllist=[]
        #print "sheetName.iter_rows():", sheetName.iter_rows()
        for row in sheetName.rows:
            r=[]
            for cell in row:
                r.append(cell.value)
                #print "cell.value", cell.value
            #print "r:", r
            fulllist.append(r)

        listwewant=[(unicode('H'),((removepunctuation(removebrtag(x[0])).lower()),removepunctuation(removebrtag(unicode(x[1]))).lower(),(x[2],x[3],x[4],x[6],x[7],len(removepunctuation(removebrtag(x[0]))),len(removepunctuation(removebrtag(unicode(x[1]))))))) if x[-1]>=0.65 else (unicode('NH'),(removepunctuation(removebrtag(x[0])).lower(),removepunctuation(removebrtag(unicode(x[1]))).lower(),(x[2],x[3],x[4],x[6],x[7],len(removepunctuation(removebrtag(x[0]))),len(removepunctuation(removebrtag(unicode(x[1])))))))for x in fulllist]
        print "hey"
        #print "listwewant", (listwewant)
    #addlist=[[x[2],x[3],x[4],x[6],x[7]]for x in fulllist]
    #listwewant=shuffle(listwewant)

    #listwewant=listwewant[:100]
        print(len(listwewant))
        samplesize=min([100,len(listwewant)])
        listwewant=sample(listwewant,samplesize)
        listwetrulywant=listwetrulywant+listwewant
    samplesize=min([250,len(listwetrulywant)])
    listwetrulywant=sample(listwetrulywant,samplesize)
    print(listwetrulywant)
    return listwetrulywant
#####################################
def removepunctuation(s):

    s = "".join(l for l in s if l not in string.punctuation)

    return s

def removebrtag(s):
    """
    """
    seed='<br/>'
    #print type(s)
    if s:
        cleaners= re.sub(seed, " ", s)

        cleaners=re.sub("\s+", " ", cleaners)
    #tweet=tweet.trim()
        cleaners=cleaners.rstrip()
        cleaners=cleaners.lstrip()
    else:
        cleaners=""
    return cleaners


def getThreeColumnDataDict(emotionLines):
    #shuffle(emotionLines)
    emotionLines=emotionLines
    classes= ["H", "NH"]
    myData={pair[0]: [] for pair in emotionLines}
    for cat in classes:
        for pair in emotionLines:
            if pair[0]==cat:
                myData[pair[0]].append(pair[1])
    return myData

def getDataStats(myData):
    # Print some stats:
    ##########################
    majorClass=max([len(myData[k]) for k in myData])
    totalCount=sum([len(myData[k]) for k in myData])
    print "Majority class count: ", majorClass
    print "Total data point count: ", totalCount
    print "Majority class % in train data: ", round((majorClass/float(totalCount))*100, 2), "%"
    print "*"*50, "\n"

def getLabeledDataTuples(myData):
    # At this point "myData" is a dict, with each emotion class as a key, and related tweet lines as a list of lines
    ###############################################################
    # The below gets me tweet body only (and filters out rest of each tweet line [e.g., tweetId.])
    # newData will be a list of tuples, each tuple has 0 as an emotion class and 1 as the string/unicode of the tweet body
    dataTuples=[(k, "".join(myData[k][i]).split("\t")[-1]) for k in myData for i in range(len(myData[k]))]
    #shuffle(dataTuples)
    #######################################################################
    # See it:
    #print "The type of newData[0][0] is a: ", type(newData[0][0]), newData[0][0] # --> newData[0] is a string
    #print "The type of newData[0][1] is a: ", type(newData[0][1]), newData[0][1] # --> newData[1] is a unicode of tweet body
    #######################################################################
    return dataTuples

def getFeatures(dataPoint):
    features=defaultdict()
    # label is class name, of course, and feats is just a list of words in this case.
    label, feats=dataPoint[0], dataPoint[1].split()
    # I could also add some code to remove the seeds from the feature dict instead of the heavy computation in
    # the tweet cleaning in removeSeed
    ###########################################
    # Beautify the below, building "has(word): True/False" dict
    for i in feats:
        features[i]=i
    if "#fearful" in features:
        del features["#fearful"]
    if "#scared" in features:
        del features["#scared"]
    return features, label

#featuresets=[getFeatures(i) for i in newData]

def getLabelsAndVectors(dataTuples):
    """
    Input:
        dataTuples is a list of tuples
        Each tuple in the list has
                   0=label
                   1= tweet body as unicode/string
    Returns an array of labels and another array for words
    """
    labels=[]
    vectors=[]
    ids=[]
    c=0
    #unicodedata.normalize('NFKD', title).encode('ascii','ignore')
    for dataPoint in dataTuples:
        ids.append(c)
        c+=1
        label, vector=dataPoint[0], (dataPoint[1][0].split(),dataPoint[1][1].split(),dataPoint[1][2])
        #label, vector=dataPoint[0], unicodedata.normalize('NFKD', dataPoint[1]).encode('ascii','ignore').split()
        labels.append(label)
        vectors.append(vector)
    #labels=array(labels)
    #print labels.shape
    #vectors=array(vectors)
    #print vectors.shape
    return ids, labels, vectors

def getSpace(vectors):
    # get the dictionary of all words in train; we call it the space as it is the space of features for bag of words
    space={}
    for dataPoint in vectors:
        words=dataPoint[0]
        for w in words:
            if w not in space:
                space[w]=len(space)
        titlewords=dataPoint[1]
        for w in titlewords:
            if "Title"+w not in space:
                space["Title"+w]=len(space)
    return space

def augmentSpace(space, featuresList):
    """
    Adds a list of features to the bag-of-words dictionary, we named "space".
    """
    for f in featuresList:
        if f not in space:
            space[f]=len(space)
    return space

def getReducedSpace(vectors, space):
    # get the dictionary of all words in train; we call it the space as it is the space of features for bag of words
    reducedSpace=defaultdict(int)
    for dataPoint in vectors:
        words=dataPoint
        for w in words:
            reducedSpace[w]+=1
    for w in space:
        if reducedSpace[w] < 3:
            del reducedSpace[w]
    reducedSpace={w: reducedSpace[w] for w in reducedSpace}
    return reducedSpace

#-------------------------------------
def addaddfeatures(vec, addf, space):
    """
    """''''
    anger, anticipation, disgust, fear, happiness, negative, positive, sadness, surprise, trust=getNRCLexicon()
    for w in sent:
        if w in anger:
            vec[space["hasAngerWord"]]=1
        if w in disgust:
            vec[space["hasDisgustWord"]]=1
        if w in fear:
            vec[space["hasFearWord"]]=1
        if w in happiness:
            vec[space["hasHappinessWord"]]=1
        if w in sadness
            vec[space["hasSadnessWord"]]=1
        if w in surprise:
            vec[space["hasSurpriseWord"]]=1
            '''
    #    AddFeatures=["STARS","VERIFIEDPURCHASE","BADGE","COMMENTS","FORMAT"]
    #    AddFeatures=["TEXTLEN","TITLELEN","STARS","VERIFIEDPURCHASE","BADGE","COMMENTS","FORMAT"]
    vec[space["TEXTLEN"]]=int(addf[-2])
    vec[space["TITLELEN"]]=int(addf[-1])
    vec[space["STARS"]]=int(addf[0])
    #badges=["null","#1 REVIEWER","TOP 10 REVIEWER","TOP 50 REVIEWER","TOP 500 REVIEWER","TOP 1000 REVIEWER","HALL OF FAMEAGEREVIEWER","THE","AMAZONOFFICIAL","AUTHOR","ARTIST","MANUFACTURER","VINE VOICE","2008 HOLIDAY TEAM","COMMUNITY FORUM 04"]
    #badgenumbers={}
    #for b in badges:
    #    badgenumbers[badges]=len(badgenumbers)
    if addf[2]!="null":
        vec[space["BADGE"]]=1

    vec[space["VERIFIEDPURCHASE"]]=int(addf[1])
    #vec[space["BADGE"]]=int(addf[2])
    vec[space["COMMENTS"]]=int(addf[3])
    formats=["Unidentified","Paperback","Kindle Edition","Hardcover","Audible Audio Edition","Audio CD"]
    phoneformats=["Unidentified","16 GB","32 GB","64 GB"]
    laptopformat=["Unidentified","ASUS Laptop","Personal Computers"]
    vec[space["FORMAT"]]=laptopformat.index(addf[4])
    return vec





#-------------------------------------------------
def getOneHotVectors(ids, labels, vectors, space):
    oneHotVectors={}
    triples=zip(ids, labels, vectors)
    #vec = np.zeros((len(space)))
    #for dataPoint in vectors:
    for triple in triples:
        idd, label, dataPoint= triple[0], triple[1], triple[2]
        #for t in xrange(len(space)):
        # populate a one-dimensional array of zeros of shape/length= len(space)
        vec=np.zeros((len(space))) # ; second argument is domensionality of the array, which is 1
        for w in dataPoint[0]:

            try:
                vec[space[w]]=1
            except:
                continue
        for w in dataPoint[1]:

            try:
                vec["Title"+space[w]]=1
            except:
                continue
        # add emotion lexicon features
        vec=addaddfeatures(vec, dataPoint[2], space)

        oneHotVectors[idd]=(vec, array(label))
    return oneHotVectors

def getOneHotVectorsAndLabels(oneHotVectorsDict):
    vectors= array([oneHotVectorsDict[k][0] for k in oneHotVectorsDict])
    labels= array([oneHotVectorsDict[k][1] for k in oneHotVectorsDict])
    print "labels.shape", labels.shape
    print "vectors.shape", vectors.shape
    return vectors, labels
###############################
# try:
#     vectors.shape[0]
# except:
#     vectors=zeros(len(vectors))

# Do grid search
#######################################
def SVM_gridSearch(trainVectors, trainLabels, kernel):
    C_range = 10.0 ** arange(-2, 2)
    gamma_range = 10.0 ** arange(-2, 2)
    param_grid = dict(gamma=gamma_range, C=C_range)
    cv = StratifiedKFold(y=trainLabels, n_folds=2)
    grid = GridSearchCV(SVC(kernel=kernel), param_grid=param_grid, cv=cv, n_jobs=n_jobs) #GridSearchCV(SVC(kernel=kernel, class_weight='auto')
    grid.fit(trainVectors, trainLabels)
    ##################################
    ## Estimated best parameters
    C = grid.best_estimator_.C
    gamma = grid.best_estimator_.gamma
    ##################################
    return C, gamma
#######################################

def getCAndGamma(trainVectors, trainLabels, kernel = 'rbf'):
    C, gamma = SVM_gridSearch(trainVectors, trainLabels, kernel)
    print C
    print gamma
    return C, gamma



def main():
    #######################################
    # Saima Aman emotion blog data
    # replacing with our data
    global AddFeatures
    AddFeatures=["TEXTLEN","TITLELEN","STARS","VERIFIEDPURCHASE","BADGE","COMMENTS","FORMAT"]
    ourdatatuples = getmyxls()
    print "Length of ourdatatuples is: ",  len(ourdatatuples)
    #shuffle(saimaDataTuples)
    print "saimaDataTuples", ourdatatuples[0]
    trainTuples=ourdatatuples#[:1000]
    #testTuples=saimaDataTuples[1000:]

#     #######################################
    myData=getThreeColumnDataDict(ourdatatuples)
    #print(myData)
    print "lol: mydata "
    #print(myData)
    totalCount=sum([len(myData[k]) for k in myData])
    print totalCount
#     del trainLines
#     print"*"*50
    getDataStats(myData)
#     dataTuples=getLabeledDataTuples(myData)
#     ####################################
#     # Add first 1000 Saima tuples
#     #dataTuples=dataTuples+saimaDataTuples[:1000]
#     print dataTuples[0]
#     del myData
    ids, labels, vectors= getLabelsAndVectors(trainTuples)
    #print labels
    space=getSpace(vectors)
    print "Total # of features in your space is: ", len(space)
    # augment space with emotion features...

    space= augmentSpace(space, AddFeatures)
    #reducedSpace=getReducedSpace(vectors, space)
    print "Total # of features in your augmented space is: ", len(space)
    print "Predicted error"
    #print "Total # of features in your reducedSpace is: ", len(reducedSpace)
    oneHotVectors=getOneHotVectors(ids, labels, vectors, space)
    print(oneHotVectors[0])
    vectors, labels=getOneHotVectorsAndLabels(oneHotVectors)
    del oneHotVectors
    trainVectors = vectors
    trainLabels = labels
    #trainLabels.fit_transform([('H','NH')])
    #trainLabels = preprocessing.label_binarize(trainLabels,classes=[unicode("H"),unicode("NH")])

    #del vectors
    #del labels
    #C, gamma = getCAndGamma(trainVectors, trainLabels, kernel = 'rbf')
    # Train classifier

    clf = OneVsOneClassifier(SVC(kernel='linear', class_weight='auto', verbose= True, probability=True))
    #clf = OneVsRestClassifier(SVC(C=1, kernel = 'linear', gamma=1, verbose= False, probability=False))

    clf.fit(trainVectors, trainLabels)
    print "\nDone fitting classifier on training data...\n"
    #testVectors = vectors[200:250]
    #testLabels = labels[200:250]
    #predicted_testLabels = clf.predict(testVectors)
    #print "Done predicting on DEV data...\n"
    #print "classification_report:\n", classification_report(testLabels, predicted_testLabels)#, target_names=target_names)
    #print "accuracy_score:", round(accuracy_score(testLabels, predicted_testLabels), 2)

    #del trainVectors
    #del trainLabels
#     saimaDataTuples=getSAIMAThreeColumnFormat()
#     print "Length of saimaDataTuples is: ",  len(saimaDataTuples)
#     shuffle(saimaDataTuples)
#     print "saimaDataTuples", saimaDataTuples[0]
#     ids, labels, vectors= getLabelsAndVectors(testTuples)
#     oneHotVectors=getOneHotVectors(ids, labels, vectors, space)
#     vectors, labels=getOneHotVectorsAndLabels(oneHotVectors)
#     del oneHotVectors
#     testVectors = vectors
#     testLabels = labels
#     predicted_testLabels = clf.predict(testVectors)




    #------------------------------------------------------------------------------------------
    print "="*50, "\n"
    print "Results with 5-fold cross validation:\n"
    print "="*50, "\n"
    #------------------------------------------------------------------------------------------
    predicted = cross_validation.cross_val_predict(clf, trainVectors, trainLabels, cv=5)
    print "*"*20
    print "\t accuracy_score\t", metrics.accuracy_score(trainLabels, predicted)
    print "*"*20

    print "precision_score\t", metrics.precision_score(trainLabels, predicted,pos_label=unicode("H"),average='binary')
    print "recall_score\t", metrics.recall_score(trainLabels, predicted,pos_label=unicode("H"),average='binary')

    print "\nclassification_report:\n\n", metrics.classification_report(trainLabels, predicted)
    print "\nconfusion_matrix:\n\n", metrics.confusion_matrix(trainLabels, predicted)

    '''#"------------------------------------------------------------------------------------------
    print "="*50, "\n"
    print "Results with 10-fold cross validation:\n"
    print "="*50, "\n"
    #------------------------------------------------------------------------------------------
    predicted = cross_validation.cross_val_predict(clf, trainVectors, trainLabels, cv=10)
    print "*"*20
    print "\t accuracy_score\t", metrics.accuracy_score(trainLabels, predicted)
    print "*"*20

    print "precision_score\t", metrics.precision_score(trainLabels, predicted,pos_label=unicode("H"),average='binary')
    print "recall_score\t", metrics.recall_score(trainLabels, predicted,pos_label=unicode("H"),average='binary')

    print "\nclassification_report:\n\n", metrics.classification_report(trainLabels, predicted)
    print "\nconfusion_matrix:\n\n", metrics.confusion_matrix(trainLabels, predicted)

    #------------------------------------------------------------------------------------------
    # Take a look at the metrics module at: http://scikit-learn.org/stable/modules/classes.html#module-sklearn.metrics
    #------------------------------------------------------------------------------------------
'''
if __name__ == "__main__":
    print "Hello!!"

    main()
    #print "Hello from the other side!!"
